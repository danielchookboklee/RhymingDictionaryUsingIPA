import openpyxl

newfile = open("word-frequency-list-60000-English-list.txt", "a")

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("word-frequency-list-60000-English.xlsx")
 
# Define variable to read sheet
dataframe1 = dataframe.active
 
# Iterate the loop to read the cell values
for row in range(1, dataframe1.max_row):
    for col in dataframe1.iter_cols(3,3):
        newfile.write(col[row].value.strip().lower())
        newfile.write("\n")
        print(col[row].value.strip().lower())


